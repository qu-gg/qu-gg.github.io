#!/usr/bin/env python3
"""
Converts the Adversaries sheet from the Unofficial BREAK!! Homebrew Adversary Database XLSX
into a JSON file for the web viewer.

Usage:
    python scripts/build_adversary_database.py <path_to_xlsx>

Output:
    break-homebrew-adversary-compendium.json in the repo root
"""
import json
import re
import sys
import os
from datetime import datetime, date
from html import escape

try:
    import openpyxl
    from openpyxl.cell.rich_text import CellRichText, TextBlock
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Columns that should be numeric stats (datetime values here are Excel m/d misinterpretation)
NUMERIC_COLS = {'ATK', 'DEF', 'HP', 'Rank', 'M', 'D', 'G', 'I', 'A'}


# Columns that should preserve rich text formatting as HTML
RICH_TEXT_COLS = {
    'Flavor Text', 'Ability 1', 'Ability 2', 'Ability 3', 'Ability 4',
    'Ability 5', 'Ability 6', 'Ability 7', 'Yield 1', 'Yield 2',
    'Tactics', 'Combat Gear', 'Habitat', 'Communication',
    'Indicators', 'Role Playing Notes', 'Customization'
}


def rich_text_to_html(cell_value):
    """Convert an openpyxl CellRichText value to an HTML string preserving bold/italic/underline."""
    if not isinstance(cell_value, CellRichText):
        text = str(cell_value) if cell_value else ""
        return escape(text).replace('\n', '<br>')

    parts = []
    for part in cell_value:
        if isinstance(part, TextBlock):
            text = escape(str(part.text)) if part.text else ""
            text = text.replace('\n', '<br>')
            font = part.font
            if font:
                if font.b:
                    text = f'<b>{text}</b>'
                if font.i:
                    text = f'<i>{text}</i>'
                if font.u and font.u != 'none':
                    text = f'<u>{text}</u>'
            parts.append(text)
        else:
            text = escape(str(part)).replace('\n', '<br>')
            parts.append(text)

    return ''.join(parts)


def convert_dash_lists(html):
    """Replace '- ' list items with bullet points for cleaner presentation."""
    # Pattern: <br> or closing tags (</b>, </i>, </u>) followed by optional whitespace and '- '
    html = re.sub(r'(<br>|</[biu]>)\s*- ', r'\1 • ', html)
    # Also handle leading '- ' at the very start of the string
    if html.startswith('- '):
        html = '• ' + html[2:]
    return html


def normalize_bold_titles(html):
    """Ensure the first bold tag (ability title) always has a line break after it.
    Subsequent bold tags (inline labels like Success/Failure) are left as-is."""
    # Case 1: <b>Title.<br></b> -> <b>Title.</b><br>  (move <br> outside)
    html = re.sub(r'<b>(.*?)<br></b>', r'<b>\1</b><br>', html)
    # Trim trailing whitespace inside bold tags (after moving <br> out)
    html = re.sub(r'<b>(.*?)\s+</b>', r'<b>\1</b>', html)
    # Remove trailing period from bold ability titles (first bold only)
    html = re.sub(r'^<b>(.*?)\.</b>', r'<b>\1</b>', html)
    # Case 2: <b>Title.</b> <br> -> <b>Title.</b><br>  (remove extra space)
    html = re.sub(r'</b>\s+<br>', '</b><br>', html)
    # Case 3: Only for the FIRST bold tag — add <br> if not already present
    def add_br_to_first_bold(m):
        return m.group(0) if m.group(0).endswith('<br>') else m.group(0) + '<br>'
    html = re.sub(r'^<b>.*?</b>(?:<br>)?', add_br_to_first_bold, html)
    return html


# Credit name normalization map (minority form -> majority form)
CREDIT_FIXES = {
    'victorseven [Discord]': 'Victorseven [Discord]',
}


def clean_entry(entry):
    """Apply presentation-layer cleanups to a single adversary entry."""
    # Strip whitespace and collapse newlines in short text fields
    for key in ('Adversary', 'Credit', 'Type', 'SPD', 'Size', 'Allegiance', 'Weapon/Shield'):
        val = entry.get(key, '')
        if isinstance(val, str):
            val = val.strip()
            if key in ('Adversary', 'Credit'):
                val = val.replace('\n', ' ').strip()
            # Collapse multiple spaces to single space
            val = re.sub(r'  +', ' ', val)
            entry[key] = val

    # Normalize Dual-Wield casing in Weapon/Shield
    w = entry.get('Weapon/Shield', '')
    if isinstance(w, str):
        w = re.sub(r'[Dd]ual-[Ww]ield(?:ing)?', 'Dual-Wield', w)
        entry['Weapon/Shield'] = w

    # Normalize credit names
    credit = entry.get('Credit', '')
    if credit in CREDIT_FIXES:
        entry['Credit'] = CREDIT_FIXES[credit]

    # Compact ability slots: shift non-empty abilities left to fill gaps
    ability_keys = [f'Ability {i}' for i in range(1, 8)]
    filled = [entry[k] for k in ability_keys if entry.get(k, '').strip()]
    for i, key in enumerate(ability_keys):
        entry[key] = filled[i] if i < len(filled) else ''

    return entry


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, rich_text=True)
    ws = wb["Adversaries"]

    # Row 1 is the banner/metadata row - extract last updated date
    last_updated = None
    for cell in ws[1]:
        val = cell.value
        if isinstance(val, CellRichText):
            val = str(val)
        if val and "Last Updated" in str(val):
            last_updated = str(val).replace("Last Updated: ", "").strip()
            break

    # Row 2 is the header row
    headers = []
    for cell in ws[2]:
        val = cell.value
        if isinstance(val, CellRichText):
            val = str(val)
        headers.append(str(val) if val else None)

    # Strip trailing None headers
    while headers and headers[-1] is None:
        headers.pop()

    # Build adversary entries from row 3 onward
    adversaries = []
    for row in ws.iter_rows(min_row=3, max_col=len(headers)):
        # Skip empty rows (no adversary name)
        first_val = row[0].value
        if isinstance(first_val, CellRichText):
            first_val = str(first_val)
        if not first_val:
            continue

        entry = {}
        for i, header in enumerate(headers):
            if header is None:
                continue
            cell = row[i] if i < len(row) else None
            val = cell.value if cell else None

            if val is None:
                entry[header] = ""
            elif header in RICH_TEXT_COLS:
                html = rich_text_to_html(val)
                html = normalize_bold_titles(html)
                html = convert_dash_lists(html)
                entry[header] = html
            elif isinstance(val, CellRichText):
                entry[header] = str(val)
            elif isinstance(val, (datetime, date)) and header in NUMERIC_COLS:
                # Excel misinterpreted a number like "12/14" as a date (m/d format).
                # The actual stat value is the month component.
                entry[header] = val.month
                print(f"  Warning: Fixed date-as-number in '{header}' for row {first_val}: {val} -> {val.month}")
            elif isinstance(val, (datetime, date)):
                entry[header] = val.isoformat()
            elif isinstance(val, float) and val == int(val):
                entry[header] = int(val)
            else:
                entry[header] = val
        adversaries.append(entry)

    # Apply presentation cleanups
    adversaries = [clean_entry(a) for a in adversaries]

    output = {
        "lastUpdated": last_updated,
        "count": len(adversaries),
        "headers": [h for h in headers if h is not None],
        "adversaries": adversaries,
    }

    out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "break-homebrew-adversary-compendium.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(adversaries)} adversaries to {out_path}")
    print(f"Last updated: {last_updated}")


if __name__ == "__main__":
    main()
