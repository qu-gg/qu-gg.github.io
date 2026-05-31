#!/usr/bin/env python3
"""
Converts the Items sheet from the Unofficial BREAK!! Homebrew Adversary Database XLSX
into a JSON file for the web viewer.

The Items sheet is transposed: each item is a column, fields are rows.
Row 1: Item name, Row 2: Credit, Row 3: Type, Row 4: Flavor Text, Row 5: Description,
Row 6: Slots, Row 7: Cost.

Usage:
    python scripts/build_item_compendium.py <path_to_xlsx>

Output:
    break-homebrew-item-compendium.json in the repo root
"""
import json
import re
import sys
import os
from html import escape

try:
    import openpyxl
    from openpyxl.cell.rich_text import CellRichText, TextBlock
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Row indices (1-based) and their field names
FIELD_ROWS = {
    1: 'Item',
    2: 'Credit',
    3: 'Type',
    4: 'Flavor Text',
    5: 'Description',
    6: 'Slots',
    7: 'Cost',
}

RICH_TEXT_FIELDS = {'Flavor Text', 'Description'}


def rich_text_to_html(cell_value):
    """Convert an openpyxl CellRichText value to an HTML string preserving bold/italic/underline."""
    if not isinstance(cell_value, CellRichText):
        text = str(cell_value) if cell_value else ""
        return escape(text).replace('\n', '<br>')

    parts = []
    for part in cell_value:
        if isinstance(part, TextBlock):
            text = escape(str(part.text)) if part.text else ""
            text = text.replace('\n', '<br>')
            font = part.font
            if font:
                if font.b:
                    text = f'<b>{text}</b>'
                if font.i:
                    text = f'<i>{text}</i>'
                if font.u and font.u != 'none':
                    text = f'<u>{text}</u>'
            parts.append(text)
        else:
            text = escape(str(part)).replace('\n', '<br>')
            parts.append(text)

    return ''.join(parts)


def convert_dash_lists(html):
    """Replace '- ' list items with bullet points for cleaner presentation."""
    html = re.sub(r'(<br>|</[biu]>)\s*- ', r'\1 • ', html)
    if html.startswith('- '):
        html = '• ' + html[2:]
    return html


def clean_item(item):
    """Apply presentation-layer cleanups to a single item entry."""
    for key in ('Item', 'Credit', 'Type'):
        val = item.get(key, '')
        if isinstance(val, str):
            val = val.strip().replace('\n', ' ')
            val = re.sub(r'  +', ' ', val)
            item[key] = val

    # Normalize Slots: convert float-like strings, strip whitespace
    slots = item.get('Slots', '')
    if isinstance(slots, float) and slots == int(slots):
        item['Slots'] = str(int(slots))
    elif isinstance(slots, float):
        item['Slots'] = str(slots)
    elif isinstance(slots, str):
        item['Slots'] = slots.strip()

    # Clean Cost
    cost = item.get('Cost', '')
    if isinstance(cost, str):
        item['Cost'] = cost.strip()

    return item


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, rich_text=True)
    ws = wb["Items"]

    # Determine how many item columns exist (starting at col 2)
    max_col = ws.max_column
    items = []

    for col_idx in range(2, max_col + 1):
        # Check if this column has an item name
        name_cell = ws.cell(row=1, column=col_idx).value
        if isinstance(name_cell, CellRichText):
            name_cell = str(name_cell)
        if not name_cell or not str(name_cell).strip():
            continue

        item = {}
        for row_idx, field_name in FIELD_ROWS.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value

            if val is None:
                item[field_name] = ''
            elif field_name in RICH_TEXT_FIELDS:
                html = rich_text_to_html(val)
                html = convert_dash_lists(html)
                item[field_name] = html
            elif isinstance(val, CellRichText):
                item[field_name] = str(val)
            elif isinstance(val, float) and val == int(val):
                item[field_name] = int(val)
            else:
                item[field_name] = val

        items.append(clean_item(item))

    # Extract last updated from the Adversaries sheet banner (shared metadata)
    last_updated = None
    try:
        ws_adv = wb["Adversaries"]
        for cell in ws_adv[1]:
            val = cell.value
            if isinstance(val, CellRichText):
                val = str(val)
            if val and "Last Updated" in str(val):
                last_updated = str(val).replace("Last Updated: ", "").strip()
                break
    except (KeyError, Exception):
        pass

    output = {
        "lastUpdated": last_updated,
        "count": len(items),
        "items": items,
    }

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "break-homebrew-item-compendium.json"
    )
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(items)} items to {out_path}")
    print(f"Last updated: {last_updated}")


if __name__ == "__main__":
    main()
